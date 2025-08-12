import os, traceback, mimetypes
from io import BytesIO
from flask import Flask, request, jsonify, send_file, Response
from PIL import Image
from werkzeug.utils import secure_filename
import zipfile, requests
from add_logo_util import paste_logo

app = Flask(__name__)

# Resolve logos path robustly under Vercel function bundle
BASES = [
    os.getcwd(),
    os.path.join(os.path.dirname(__file__), ".."),
    os.path.dirname(__file__),
]
def _resolve_logos_root():
    for base in BASES:
        cand = os.path.abspath(os.path.join(base, 'logos'))
        if os.path.isdir(cand):
            return cand
    # Fallback: relative to this file
    return os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'logos'))
LOGOS_ROOT = _resolve_logos_root()

def find_logo_for_language(language: str):
    """
    Prefer *_RGB.png then any .png then .jpg/.jpeg under logos/<language>/**.
    """
    if not language:
        raise ValueError("language is required")
    lang_folder = os.path.join(LOGOS_ROOT, language)
    if not os.path.isdir(lang_folder):
        # Try case-insensitive match of folder name
        entries = [d for d in os.listdir(LOGOS_ROOT) if os.path.isdir(os.path.join(LOGOS_ROOT, d))]
        for d in entries:
            if d.lower() == language.lower():
                lang_folder = os.path.join(LOGOS_ROOT, d)
                break
    if not os.path.isdir(lang_folder):
        raise FileNotFoundError(f"Language folder not found: {language}")

    preferred_png = None
    any_png = None
    any_jpg = None
    for root, dirs, files in os.walk(lang_folder):
        for name in files:
            lower = name.lower()
            full = os.path.join(root, name)
            if lower.endswith("_rgb.png") and preferred_png is None:
                preferred_png = full
            elif lower.endswith(".png") and any_png is None:
                any_png = full
            elif lower.endswith((".jpg", ".jpeg")) and any_jpg is None:
                any_jpg = full
    path = preferred_png or any_png or any_jpg
    if not path:
        raise FileNotFoundError(f"No logo image found under: {lang_folder}")
    return Image.open(path)

def _read_image_from_request_file(fs):
    data = fs.read()
    img = Image.open(BytesIO(data))
    # Force load now to avoid lazy file handles issues in serverless
    img.load()
    return img, fs.filename or "upload.jpg"

def _read_image_from_url(url: str):
    r = requests.get(url, timeout=30)
    r.raise_for_status()
    img = Image.open(BytesIO(r.content))
    img.load()
    # try to derive a safe filename
    base = os.path.basename(url.split("?")[0]) or "remote.jpg"
    return img, base

def _safe_output_name(base: str, language: str, to_ext=None):
    root, ext = os.path.splitext(secure_filename(base))
    ext = to_ext or (ext if ext else ".jpg")
    return f"{root}_{language}{ext}"

def _parse_logo_scale(raw: str):
    try:
        v = float(raw)
    except Exception:
        v = 0.40
    return max(0.05, min(1.0, v))

@app.post("/")
@app.post("/api/add_logo")
@app.post("/add_logo")
@app.post("/api/add_logo")
def add_logo():
    try:
        language = request.form.get("language")
        logo_scale = _parse_logo_scale(request.form.get("logo_scale", "0.40"))
        if not language:
            return jsonify({"error": "language is required"}), 400

        logo_img = find_logo_for_language(language)

        # Collect images: files
        images_to_process = []
        for key, storages in request.files.lists():
            if key == "excel":
                continue
            for fs in storages:
                try:
                    img, fname = _read_image_from_request_file(fs)
                    images_to_process.append((img, fname))
                except Exception:
                    continue

        # URLs
        for u in request.form.getlist("url"):
            u = (u or "").strip()
            if not u:
                continue
            try:
                img, fname = _read_image_from_url(u)
                images_to_process.append((img, fname))
            except Exception:
                continue

        # Excel (optional)
        if "excel" in request.files:
            try:
                from openpyxl import load_workbook
                wb = load_workbook(request.files["excel"], read_only=True, data_only=True)
                ws = wb.active
                # Heuristic: first row headers "url" or "image_url", else take first column
                url_col_idx = 1
                first_row = next(ws.iter_rows(min_row=1, max_row=1))
                headers = [(c.value or "") if c.value is not None else "" for c in first_row]
                headers_norm = [str(h).strip().lower() for h in headers]
                for idx, h in enumerate(headers_norm, start=1):
                    if h in ("url", "image_url"):
                        url_col_idx = idx
                        break
                for row in ws.iter_rows(min_row=2 if any(headers_norm) else 1):
                    val = row[url_col_idx-1].value
                    if not val:
                        continue
                    u = str(val).strip()
                    if not u:
                        continue
                    try:
                        img, fname = _read_image_from_url(u)
                        images_to_process.append((img, fname))
                    except Exception:
                        continue
            except Exception:
                pass

        if not images_to_process:
            return jsonify({"error": "no images provided (files, urls, or excel)"}), 400

        outputs = []
        for img, fname in images_to_process:
            composed = paste_logo(img, logo_img, logo_scale=logo_scale)
            out = BytesIO()
            ext = (os.path.splitext(fname)[1] or ".jpg").lower()
            if ext == ".png":
                composed.save(out, format="PNG", optimize=True)
                out.seek(0)
                outputs.append((_safe_output_name(fname, language, ".png"), "image/png", out))
            else:
                composed.convert("RGB").save(out, format="JPEG", quality=100, optimize=True)
                out.seek(0)
                outputs.append((_safe_output_name(fname, language, ".jpg"), "image/jpeg", out))

        if len(outputs) == 1:
            name, mime, bio = outputs[0]
            return send_file(bio, mimetype=mime, as_attachment=True, download_name=name)

        # multiple → zip
        zip_buf = BytesIO()
        with zipfile.ZipFile(zip_buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            for name, mime, bio in outputs:
                zf.writestr(name, bio.getvalue())
        zip_buf.seek(0)
        return send_file(zip_buf, mimetype="application/zip", as_attachment=True, download_name="logo_stamped_images.zip")

    except Exception as e:
        tb = traceback.format_exc()
        print("FUNCTION ERROR", e, tb, flush=True)
        return Response(f"Error: {e}\n\n{tb}", status=500, mimetype="text/plain")
